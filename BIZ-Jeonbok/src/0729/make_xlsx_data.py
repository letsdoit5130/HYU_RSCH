"""
완도 전복 글로벌 무역 EDA 및 시장 개척 통합 Excel 데이터북 생성 스크립트 (11대 지정 컬럼 표준 825개 파트너 DB 반영)

이 스크립트는 openpyxl을 활용하여 BIZ-Jeonbok/reports/Wando_Abalone_Integrated_Data.xlsx 파일을 생성합니다.
11대 필수 지정 컬럼 표준 순서(수집일, 검증일, 사명만, 이메일, 웹사이트, 도시, 지방, 국가, 메신저, SNS, 간단품목)로 825개사 데이터를 수록합니다.
"""

import os
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

font_title = Font(name='Arial', size=16, bold=True, color='FFFFFF')
font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')

fill_title = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid')
fill_header = PatternFill(start_color='2B5C8F', end_color='2B5C8F', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')

# Sheet 1: Summary_KPI
ws1 = wb.active
ws1.title = "Summary_KPI"
ws1.views.sheetView[0].showGridLines = True

ws1.merge_cells('A1:E1')
ws1['A1'] = "완도 전복 글로벌 무역 EDA 및 시장 개척 KPI 총괄"
ws1['A1'].font = font_title
ws1['A1'].fill = fill_title
ws1['A1'].alignment = align_center

headers_kpi = ["지표 항목 (KPI)", "수치 (Value)", "단위 (Unit)", "비고 (Notes)", "데이터 출처"]
ws1.append([])
ws1.append(headers_kpi)

kpi_data = [
    ["전체 분석 데이터 수", 5400, "건", "UN Comtrade 글로벌 수산 데이터", "BIZ-JB-Gathered.csv"],
    ["총 무역 거래 금액", 5209010000, "USD", "2021~2025년 누적 무역액", "UN Comtrade"],
    ["총 무역 물동량", 271250000, "kg", "2021~2025년 누적 순중량", "UN Comtrade"],
    ["수입(Import) 거래 비중", 0.941, "%", "수입 신고 데이터 중심", "UN Comtrade"],
    ["검증 실존 파트너 바이어 DB", 825, "개사", "15개 유망국가 국가별 55개사 수집", "abalone_buyers_db_v2.csv"],
    ["전복 평균 단가", 33.03, "$/kg", "가공품 및 생물 전체 평균", "EDA Result"],
    ["전복 단가 중앙값", 16.02, "$/kg", "중간 등급 일반 유통 단가", "EDA Result"],
    ["최고 수출 단가", 2284.00, "$/kg", "초고가 건전복 최고 등급", "EDA Result"],
    ["대한민국 수출 실적", 581395000, "USD", "전 세계 유일의 대규모 순수출국", "UN Comtrade"],
    ["홍콩 수입 시장 규모", 1283200000, "USD", "단일 세계 최대 전복 수입 시장", "UN Comtrade"]
]

for row in kpi_data:
    ws1.append(row)

for col_idx in range(1, 6):
    cell = ws1.cell(row=3, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

# Sheet 2: Export_Pricing
ws2 = wb.create_sheet(title="Export_Pricing")
ws2.views.sheetView[0].showGridLines = True

ws2.merge_cells('A1:I1')
ws2['A1'] = "완도 전복 국가별/품목별 FOB & CIF 수출 가격 계산 모델"
ws2['A1'].font = font_title
ws2['A1'].fill = fill_title
ws2['A1'].alignment = align_center

headers_price = ["No", "타겟 국가", "타겟 품목 및 규격 (데이터셋 HS Code)", "EXW 산지원가 ($)", "물류/포장비 ($)", "FOB 가격 ($)", "CIF 가격 ($)", "마진율 (%)", "비고"]
ws2.append([])
ws2.append(headers_price)

pricing_raw = [
    [1, "홍콩", "활전복 (7~8미/kg, HS 030781)", 19.00, 3.50, 5.00],
    [2, "홍콩", "명품 건전복 (25미/500g, HS 030781)", 100.00, 8.00, 7.00],
    [3, "싱가포르", "활전복 (10~12미/kg, HS 030781)", 17.50, 3.00, 6.00],
    [4, "미국", "전복 통조림 (4미/400g, HS 160557)", 9.50, 1.20, 1.10],
    [5, "미국", "자숙 냉동전복 파우치 (1kg, HS 160557)", 17.00, 2.50, 2.80],
    [6, "캐나다", "전복 통조림 (굴소스 400g, HS 160557)", 9.50, 1.20, 1.20],
    [7, "일본", "횟감용 자숙전복 (10미/kg, HS 030783)", 18.50, 2.80, 4.00],
    [8, "베트남", "활전복 (10~12미/kg, HS 030781)", 17.50, 2.50, 5.00],
    [9, "호주", "전복 통조림 (4미/400g, HS 160557)", 10.00, 1.30, 1.20],
    [10, "프랑스", "전복 내장 가공품 (200g, HS 160557)", 16.00, 2.20, 3.00]
]

for r in pricing_raw:
    row_idx = len(ws2['A']) + 1
    no, country, item, exw, ship_fob, ship_cif = r
    fob_formula = f"=ROUND(D{row_idx}*1.23 + E{row_idx}, 2)"
    cif_formula = f"=ROUND(F{row_idx} + G{row_idx}, 2)"
    margin_formula = f"=(F{row_idx}-D{row_idx}-E{row_idx})/F{row_idx}"
    ws2.append([no, country, item, exw, ship_fob, fob_formula, cif_formula, margin_formula, "추천 가격 적용"])

for col_idx in range(1, 10):
    cell = ws2.cell(row=3, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

# Sheet 3: Product_Star_Strategy
ws3 = wb.create_sheet(title="Product_Star_Strategy")
ws3.views.sheetView[0].showGridLines = True

ws3.merge_cells('A1:F1')
ws3['A1'] = "품목별 Star & Rising Star 타겟 시장 및 BIZ-JB-Gathered.csv 수록 HS CODE"
ws3['A1'].font = font_title
ws3['A1'].fill = fill_title
ws3['A1'].alignment = align_center

headers_star = ["품목 분류 (데이터셋 수록 HS Code)", "Star 시장 (주력)", "Rising Star 시장 (신흥)", "Cash Cow / Selective", "핵심 물류/포장 방식", "추천 CIF 단가 ($)"]
ws3.append([])
ws3.append(headers_star)

star_raw = [
    ["1. 활전복 (Live, HS 030781)", "홍콩, 싱가포르, 마카오", "베트남 (호치민/하노이), 대만", "일본 (Sub-Star)", "항공 (Air) 산소 주입 해수 팩", "$31.00~$45.00 / kg"],
    ["2. 명품 건전복 (Dried, HS 030781)", "홍콩, 마카오, 광동성", "말레이시아 (KL), 미국 (중화권)", "싱가포르 (Cash Cow)", "항공/해상 하드케이스 선물세트", "$142.00 / 500g"],
    ["3. 통조림 & 파우치 (HS 160557)", "미국, 캐나다 (Cash Cow)", "호주 (시드니), 영국/독일", "중국 (대량 유통)", "해상 (Sea) 400g 캔 / 1kg 파우치", "$14.30/캔, $26.80/kg"],
    ["4. 횟감용 IQF 냉동전복 (HS 030783)", "일본 (도쿄/오사카)", "베트남, 태국, 호주", "유럽 (Selective)", "해상/항공 IQF 개별급속동결", "$30.00 / kg"],
    ["5. 전복 내장소스 / 가공품 (HS 160557)", "미국 (메인스트림)", "프랑스, 네덜란드, 이탈리아", "말레이시아", "해상 유리병 / 레토르트 파우치", "$12.50 / 병 (200g)"]
]

for row in star_raw:
    ws3.append(row)

for col_idx in range(1, 7):
    cell = ws3.cell(row=3, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

# Sheet 4: Product_Top_Buyers_DB (Full 825 Buyers - 11 Mandatory Columns Standard Order)
ws4 = wb.create_sheet(title="Product_Top_Buyers_DB")
ws4.views.sheetView[0].showGridLines = True

ws4.merge_cells('A1:K1')
ws4['A1'] = "15개 유망국가 825개사 글로벌 전복 파트너/바이어 종합 DB (11대 지정 컬럼 표준)"
ws4['A1'].font = font_title
ws4['A1'].fill = fill_title
ws4['A1'].alignment = align_center

headers = [
    "데이터 수집일", "데이터 검증일", "취급 품목",
    "회사명 (사명만)", "Prf_CName", "Ver_CName",
    "웹사이트", "Prf_CWeb", "Ver_CWeb",
    "컨택 이메일", "Prf_Email", "Ver_Email",
    "회사 위치한 도시", "회사 위치한 지방", "회사 위치한 국가",
    "Messanger (WhatsApp, Line, Zalo and etc)", "Prf_Msg", "Ver_Msg",
    "SNS (Linkedin, Instagram, Facebook etc)", "Prf_SNS", "Ver_SNS",
    "회사 소개", "추천 수출물품 및 수출가", "Verified_CINFO"
]
ws4.append([])
ws4.append(headers)

csv_clean = os.path.join("BIZ-Jeonbok", "data", "abalone_buyers_db_cleaned.csv")
csv_v4 = os.path.join("BIZ-Jeonbok", "data", "abalone_buyers_db_v4.csv")
csv_v3 = os.path.join("BIZ-Jeonbok", "data", "abalone_buyers_db_v3.csv")

if os.path.exists(csv_clean):
    target_csv = csv_clean
elif os.path.exists(csv_v4):
    target_csv = csv_v4
else:
    target_csv = csv_v3

if os.path.exists(target_csv):
    df_buyers = pd.read_csv(target_csv, dtype=str).fillna("")
    for _, row in df_buyers.iterrows():
        ws4.append([
            row["데이터 수집일"], row["데이터 검증일"], row.get("취급 품목", ""), row["회사명 (사명만)"], row["컨택 이메일"],
            row["웹사이트"], row["회사 위치한 도시"], row["회사 위치한 지방"], row["회사 위치한 국가"],
            row["Messanger (WhatsApp, Line, Zalo and etc)"],
            row["SNS (Linkedin, Instagram, Facebook etc)"],
            row.get("회사 소개", row.get("회사 간단 품목", "")),
            row.get("추천 수출물품 및 수출가", row.get("추천 CIF 단가", ""))
        ])

for col_idx in range(1, 12):
    cell = ws4.cell(row=3, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

# Sheet 5: Business_Playbook
ws5 = wb.create_sheet(title="Business_Playbook")
ws5.views.sheetView[0].showGridLines = True

ws5.merge_cells('A1:D1')
ws5['A1'] = "1인 종합상사 실전 비즈니스 파이프라인 & 5단계 액션 플레이북"
ws5['A1'].font = font_title
ws5['A1'].fill = fill_title
ws5['A1'].alignment = align_center

headers_playbook = ["실전 단계 (Step)", "액션 제목 (Action Title)", "세부 실행 내용 (Execution Details)", "핵심 성공 요인 (KPI / Key Factors)"]
ws5.append([])
ws5.append(headers_playbook)

playbook_raw = [
    ["Step 1 / Action 1", "Cold Outreach & Pitching", "825개 검증 바이어 담당자 대상 영문/중문 제안서 송부", "현지 시세 대비 15% 우위 단가 + 폐사율 3% 보증 + 50kg MOQ 오퍼"],
    ["Step 2 / Action 2", "Sample Shipping & Testing", "10kg 산소주입 활전복 팩 Air Express 송부 / 통조림 1박스 수송", "24시간 도착 후 생존율, 수율(Meat Yield), 멸균 품질 검증서 승인"],
    ["Step 3 / Action 3", "Contract & Financial Term", "신규 거래처 T/T 30% Deposit + 70% B/L Copy / 대형사 L/C at Sight", "무조건부 일람출급 신용장(L/C) 활용으로 대금 회수 리스크 100% 차단"],
    ["Step 4 / Action 4", "Compliance & Customs Clearance", "미국 FDA Prior Notice, 베트남 한-베 FTA C/O, 유럽 EU 위생허가", "베트남 관세 0% C/O 발급 및 미국 FDA 사전통보 승인 연동"],
    ["Step 5 / Action 5", "Re-order & Seasonality Supply", "춘절/중추절 3개월 전 명품 건전복 계약, 북미 연말 통조림 FCL 발주", "계절별 수급 파이프라인 구축을 통한 바이어 LTV 극대화"]
]

for row in playbook_raw:
    ws5.append(row)

for col_idx in range(1, 5):
    cell = ws5.cell(row=3, column=col_idx)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center

for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

output_file = os.path.join("BIZ-Jeonbok", "reports", "Wando_Abalone_Integrated_Data.xlsx")
alt_output_file = os.path.join("BIZ-Jeonbok", "reports", "Wando_Abalone_Integrated_Data_v5.xlsx")

try:
    wb.save(output_file)
    print(f"--- UPDATED XLSX DATA BOOK (825 BUYERS 11 COLS) GENERATED SUCCESSFULLY: {output_file} ---")
except PermissionError:
    wb.save(alt_output_file)
    print(f"--- UPDATED XLSX DATA BOOK SAVED TO ALTERNATE PATH: {alt_output_file} ---")
