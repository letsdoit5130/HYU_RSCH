"""
BIZ-전복_Gathered_EDA_Report.md 기반 TOP 10 유망국가 표 3종 시트 완비 Excel(.xlsx) 생성 스크립트

이 스크립트는 전복 EDA 보고서의 요약 대시보드, 미수(Size) 가격 구조 매트릭스, 
HS Code 3대 품목별 (0307.81 활전복, 0307.83 냉동전복, 1605.57 전복통조림) TOP 10 유망국가 분석 시트 3종(각 10개국 완비)을 
Multi-Sheet 엑셀 파일로 변환 생성합니다.
"""
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Windows 콘솔 인코딩 방어
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX_OUTPUT = os.path.join(BASE_DIR, 'data', 'BIZ_Jeonbok_Integrated_Data.xlsx')

def create_xlsx_integrated_data():
    wb = openpyxl.Workbook()

    font_title = Font(name="맑은 고딕", size=15, bold=True, color="1F497D")
    font_sub = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    font_normal = Font(name="맑은 고딕", size=10)

    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. Sheet 1: Dashboard Summary
    ws1 = wb.active
    ws1.title = "Dashboard_Summary"
    ws1.views.sheetView[0].showGridLines = True

    ws1['A1'] = "📊 한국산 전복(Abalone) 무역 EDA 통합 데이터 대시보드"
    ws1['A1'].font = font_title

    metrics = [
        ["핵심 성과 지표 (KPI)", "측정 수치", "비고 / 1인 상사 소싱 포인트"],
        ["분석 대상 무역 레코드 수", "500건", "전 세계 전복 수출입 거래 데이터 정규화"],
        ["누적 총 무역액", "$148.50M", "전복 글로벌 수입 시장 규모"],
        ["평균 수출입 단가", "$32.40 / kg", "전체 미수(Size) 평균 단가"],
        ["최대 수입국 점유율 (일본)", "35.4%", "완도산 활전복 항공/페리 1차 도매 수입"],
        ["최대 냉동 수입국 (미국)", "42.1%", "미 서부/동부 아시안 마트 냉동 IQF 해상 수입"],
        ["최대 가공 수입국 (홍콩)", "48.5%", "홍콩 셩완 시장 명절 선물용 캔 통조림 수입"]
    ]

    for r_idx, row in enumerate(metrics, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_sub if r_idx == 3 else font_normal
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if r_idx == 3:
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif r_idx % 2 == 1:
                cell.fill = fill_zebra

    # 2. Sheet 2: Size_Pricing_Structure
    ws2 = wb.create_sheet(title="Size_Pricing_Structure")
    ws2.views.sheetView[0].showGridLines = True
    ws2['A1'] = "💰 전복 미수(Size) 및 규격별 글로벌 가격 구조 매트릭스 ($/kg)"
    ws2['A1'].font = font_title

    pricing_data = [
        ["품목 규격 / 미수", "마리당 중량 범위", "주요 수출 국가", "평균 단가 ($/kg)", "주요 타깃 시장 & 바이어 채널", "1인 상사 소싱 추천 포인트"],
        ["10미 미만 (대과)", "100g 이상 / 마리", "한국 완도, 호주, 멕시코", "$42.0 ~ $48.0", "일본 고급 일식집, 스시야, 료칸", "고급 항공직송 프리미엄 오퍼"],
        ["10 ~ 12미 (중대과)", "80g ~ 100g", "한국 완도", "$36.0 ~ $40.0", "도쿄 도요스 시장 도매상사, 미국 LA", "메인 수출 주력 미수, 1차 수입상사"],
        ["13 ~ 15미 (중과)", "65g ~ 80g", "한국, 중국", "$30.0 ~ $34.0", "관서 지역 레스토랑, 아시안 마트", "H-Mart, 99 Ranch 채널 공급"],
        ["15 ~ 20미 (중소과)", "50g ~ 65g", "한국, 베트남", "$24.0 ~ $28.0", "냉동 IQF 가공, 외식 프랜차이즈", "해상 IQF 컨테이너 대량 공급"],
        ["20미 이상 (소과/가공)", "50g 미만", "한국, 중국", "$18.0 ~ $22.0", "통조림 가공, HMR 파우치 가공", "통조림 FDA 승인 공장 연동"]
    ]

    for r_idx, row in enumerate(pricing_data, start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_sub if r_idx == 3 else font_normal
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if r_idx == 3:
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif r_idx % 2 == 1:
                cell.fill = fill_zebra

    # 3. Sheet 3~5: TOP 10 Promising Countries (각 10개국 완비!)
    sheets_info = [
        ("Top10_Fresh_Abalone", "[표 1] HS 0307.81 (활/신선 전복) TOP 10 유망 국가", [
            ["유망순위", "타깃 국가", "무역액 점유율", "컨택해야 할 로컬 파트너 종류", "1인 상사 시장개척 포인트"],
            ["1위", "일본 (Japan)", "35.4%", "도쿄 도요스 시장 수산물 수입 도매상사", "완도산 활전복 페리/항공 직송 1차 수입 도매 공급"],
            ["2위", "중국 (China)", "24.1%", "동해안 수산물 수입 및 유통 상사", "산둥성/상하이 고급 호텔 및 외식 체인 공급"],
            ["3위", "홍콩 (Hong Kong)", "18.2%", "고급 수산물 건재 시장 수입상사", "고급 딤섬 및 레스토랑 직송 공급"],
            ["4위", "대만 (Taiwan)", "7.5%", "타이베이 고급 수산물 1차 수입상", "일식 뷔페 및 연회장 활전복 대량 공급"],
            ["5위", "미국 (USA)", "4.8%", "LA/NY 아시안 수산물 벤더", "한인/아시안 고소득층 대상 항공 직송"],
            ["6위", "싱가포르 (Singapore)", "3.2%", "마리나 베이 외식 그룹 벤더", "고급 해산물 뷔페 및 호텔 공급"],
            ["7위", "베트남 (Vietnam)", "2.5%", "호치민/하노이 수산물 수입상", "한국 식당가 및 고급 수산 레스토랑"],
            ["8위", "캐나다 (Canada)", "1.8%", "밴쿠버 아시안 수산 유통사", "밴쿠버/토론토 아시안 마트 활전복"],
            ["9위", "태국 (Thailand)", "1.3%", "방콕 고급 수산물 수입 대리점", "방콕 5성급 호텔 수산물 오퍼"],
            ["10위", "호주 (Australia)", "1.2%", "시드니 아시안 식품 유통 벤더", "호주 한인 마트 및 아시안 레스토랑"]
        ]),
        ("Top10_Frozen_Abalone", "[표 2] HS 0307.83 (냉동 전복) TOP 10 유망 국가", [
            ["유망순위", "타깃 국가", "무역액 점유율", "컨택해야 할 로컬 파트너 종류", "1인 상사 시장개척 포인트"],
            ["1위", "미국 (USA)", "42.1%", "미 서부 최대 수산물 수입 벤더 (PASCO 등)", "아시안 마트향 냉동 IQF 해상 컨테이너 공급"],
            ["2위", "대만 (Taiwan)", "19.8%", "타이베이 식자재 수입 디스트리뷰터", "외식 뷔페 및 연회장향 IQF 냉동 대량 공급"],
            ["3위", "일본 (Japan)", "15.3%", "관서 지역 냉동 수산물 수입 대리점", "성수기 외식 체인 원료 공급"],
            ["4위", "홍콩 (Hong Kong)", "8.2%", "냉동 수산물 전문 수입 유통사", "외식 체인 및 호텔 냉동 IQF 공급"],
            ["5위", "싱가포르 (Singapore)", "4.5%", "동남아 아시안 식자재 유통 벤더", "뷔페 및 딤섬 프랜차이즈 공급"],
            ["6위", "중국 (China)", "3.8%", "연안 도시 식품 가공 및 유통사", "가공 원료용 냉동 IQF 전복 공급"],
            ["7위", "캐나다 (Canada)", "2.1%", "토론토 수산물 수입 벤더", "아시안 마트 냉동 해산물 코너 공급"],
            ["8위", "베트남 (Vietnam)", "1.8%", "외식 식자재 1차 수입상", "프랜차이즈 레스토랑 IQF 전복 공급"],
            ["9위", "태국 (Thailand)", "1.3%", "방콕 식자재 수입 대리점", "외식 뷔페 및 씨푸드 레스토랑 공급"],
            ["10위", "영국 (United Kingdom)", "1.1%", "런던 아시안 식품 수입 벤더", "런던 아시안 마트 및 한식당 공급"]
        ]),
        ("Top10_Canned_Abalone", "[표 3] HS 1605.57 (전복 통조림) TOP 10 유망 국가", [
            ["유망순위", "타깃 국가", "무역액 점유율", "컨택해야 할 로컬 파트너 종류", "1인 상사 시장개척 포인트"],
            ["1위", "홍콩 (Hong Kong)", "48.5%", "홍콩 셩완 수산물 건재 시장 수입상사", "춘절 명절 선물 세트용 B2B 캔 대량 공급"],
            ["2위", "싱가포르 (Singapore)", "22.1%", "싱가포르 고급 선물 세트 수입 유통 벤더", "명절/기념일 프리미엄 선물용 캔 공급"],
            ["3위", "미국 (USA)", "14.8%", "북미 아시안 식품 수입 벤더 (아씨마켓 등)", "FDA LACF 승인 캔 통조림 전역 유통"],
            ["4위", "대만 (Taiwan)", "4.2%", "명절 선물 세트 수입 유통사", "명절 고급 전복 캔 선물 세트 공급"],
            ["5위", "캐나다 (Canada)", "3.1%", "밴쿠버 아시안 마트 유통 벤더", "북미 한인/중국인 마트 캔 전복 공급"],
            ["6위", "호주 (Australia)", "2.3%", "시드니/멜버른 아시안 식품 수입상", "선물용 캔 전복 유통"],
            ["7위", "일본 (Japan)", "1.8%", "고급 통조림 식자재 유통사", "료칸 및 기프트 숍 고급 캔 오퍼"],
            ["8위", "베트남 (Vietnam)", "1.2%", "고급 선물 세트 수입상", "호치민/하노이 명절 선물용 캔 전복"],
            ["9위", "태국 (Thailand)", "1.1%", "방콕 아시안 식품 수입 벤더", "고급 아시안 마트 캔 유통"],
            ["10위", "영국 (United Kingdom)", "0.9%", "런던 프리미엄 기프트 숍 벤더", "런던 아시안 명절 기프트 공급"]
        ])
    ]

    for s_title, title_str, t_rows in sheets_info:
        ws = wb.create_sheet(title=s_title)
        ws.views.sheetView[0].showGridLines = True
        ws['A1'] = f"🗺️ {title_str}"
        ws['A1'].font = font_title

        for r_idx, row in enumerate(t_rows, start=3):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = font_sub if r_idx == 3 else font_normal
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center')
                if r_idx == 3:
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif r_idx % 2 == 1:
                    cell.fill = fill_zebra

    # Auto-fit column widths
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                max_len = max(max_len, len(val_str))
            sheet.column_dimensions[col_letter].width = max(max_len * 1.4, 12)

    wb.save(XLSX_OUTPUT)
    print(f"✅ [TOP 10 유망국가 10개국 완비] Excel(.xlsx) 데이터베이스 재생성 완료: {XLSX_OUTPUT}")

if __name__ == "__main__":
    create_xlsx_integrated_data()
