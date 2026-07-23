"""
openpyxl 기반 엑셀 요약 및 KPI 대시보드 자동 빌더 (build_excel_dashboard.py)

이 프로그램은 수집된 CSV 데이터를 바탕으로 '대시보드 요약' 시트(타이틀바, KPI 카드, 수식)와
'상세 데이터' 시트를 포함하는 비즈니스 포맷의 .xlsx 파일 대시보드를 자동 조립하는 모듈입니다.

작성일: 2026-07-23
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def build_excel_dashboard(csv_path: str = "data/raw_data.csv", output_excel: str = "data/dashboard.xlsx"):
    print(f"[EXCEL-BUILDER] 엑셀 대시보드 조립 시작 -> {output_excel}")
    if not os.path.exists(csv_path):
        print(f"[EXCEL-BUILDER ERROR] 데이터 없음: {csv_path}")
        return
        
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    wb = Workbook()
    
    # 시트 1: 대시보드 요약
    ws1 = wb.active
    ws1.title = "대시보드 요약"
    ws1["A1"] = "📊 비즈니스 수집 데이터 종합 분석 대시보드"
    ws1["A1"].font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    ws1["A1"].fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    
    ws1["A3"] = "총 수집 건수"
    ws1["B3"] = len(df)
    
    # 시트 2: 상세 데이터
    ws2 = wb.create_sheet(title="상세 데이터")
    ws2.append(list(df.columns))
    for _, row in df.iterrows():
        ws2.append(list(row))
        
    os.makedirs(os.path.dirname(output_excel), exist_ok=True)
    wb.save(output_excel)
    wb.close()
    print(f"[EXCEL-BUILDER COMPLETED] 엑셀 대시보드 저장 완수: {output_excel}")

if __name__ == "__main__":
    build_excel_dashboard()
