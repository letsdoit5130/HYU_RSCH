"""
수출 EDA 대시보드 및 통계 데이터 Excel (.xlsx) v2 파일 생성 스크립트.
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DATA_PATH = 'BIZ-laver/data/HaeYu-Laver-EXP.csv'
EXCEL_OUTPUT_PATH = 'BIZ-laver/reports/HaeYu_Laver_Export_Dashboard_and_Data_v2.xlsx'

df_raw = pd.read_csv(DATA_PATH)

def clean_curr(v):
    if pd.isna(v):
        return 0.0
    return float(str(v).replace('$', '').replace(',', '').strip())

df_raw['primary_value_usd'] = df_raw['primaryValue'].apply(clean_curr)
df_raw['unit_price_usd_kg'] = df_raw['Unit Price ($PV/kg)'].apply(clean_curr)
df_raw['qty_tons'] = df_raw['Qty (t)']

wb = openpyxl.Workbook()
ws_dash = wb.active
ws_dash.title = "Executive Dashboard"

font_title = Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF')
font_header = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
font_kpi_num = Font(name='맑은 고딕', size=14, bold=True, color='1A5274')
font_bold = Font(name='맑은 고딕', size=10, bold=True)
font_regular = Font(name='맑은 고딕', size=10)

fill_navy = PatternFill(start_color='1A5274', end_color='1A5274', fill_type='solid')
fill_blue = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
fill_light = PatternFill(start_color='F4F6F7', end_color='F4F6F7', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')

thin_border = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)

ws_raw = wb.create_sheet(title="Raw Data")
raw_headers = list(df_raw.columns)
ws_raw.append(raw_headers)

for cell in ws_raw[1]:
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_center

for row in df_raw.itertuples(index=False):
    ws_raw.append(list(row))

ws_dash.merge_cells("A1:G1")
ws_dash["A1"] = "해유 김 수출 EDA & 글로벌 마진 Executive Dashboard"
ws_dash["A1"].font = font_title
ws_dash["A1"].fill = fill_navy
ws_dash["A1"].alignment = align_center
ws_dash.row_dimensions[1].height = 40

kpis = [
    ("총 수출액 (달러)", "=SUM('Raw Data'!AE:AE)", "B3", "$#,##0"),
    ("총 수출 물량 (톤)", "=SUM('Raw Data'!AG:AG)", "D3", "#,##0.0"),
    ("평균 수출 단가 ($/kg)", "=AVERAGE('Raw Data'!AH:AH)", "F3", "$#,##0.00")
]

for title, formula, cell_ref, num_fmt in kpis:
    col_idx = openpyxl.utils.column_index_from_string(cell_ref[0])
    row_idx = int(cell_ref[1:])
    
    ws_dash.cell(row=row_idx, column=col_idx, value=title).font = font_bold
    ws_dash.cell(row=row_idx, column=col_idx).alignment = align_center
    ws_dash.cell(row=row_idx, column=col_idx).fill = fill_light
    
    val_cell = ws_dash.cell(row=row_idx+1, column=col_idx, value=formula)
    val_cell.font = font_kpi_num
    val_cell.alignment = align_center
    val_cell.number_format = num_fmt
    val_cell.border = thin_border

ws_dash.cell(row=6, column=1, value="연도별 수출 동향 (동적 수식 연결)").font = Font(name='맑은 고딕', size=12, bold=True, color='1A5274')
headers_dash = ["연도", "마른김(121221) 수출액", "조미김(200899) 수출액", "전체 합계액", "조미김 비중(%)"]

for c_idx, h in enumerate(headers_dash, start=1):
    cell = ws_dash.cell(row=7, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_center

years = [2021, 2022, 2023, 2024, 2025]
for r_offset, y in enumerate(years, start=8):
    ws_dash.cell(row=r_offset, column=1, value=y).alignment = align_center
    
    f_raw = f"=SUMIFS('Raw Data'!AE:AE, 'Raw Data'!E:E, {y}, 'Raw Data'!Z:Z, 121221)"
    f_seasoned = f"=SUMIFS('Raw Data'!AE:AE, 'Raw Data'!E:E, {y}, 'Raw Data'!Z:Z, 200899)"
    f_total = f"=SUM(B{r_offset}:C{r_offset})"
    f_ratio = f"=C{r_offset}/D{r_offset}"
    
    c_r = ws_dash.cell(row=r_offset, column=2, value=f_raw)
    c_r.number_format = "$#,##0"
    c_s = ws_dash.cell(row=r_offset, column=3, value=f_seasoned)
    c_s.number_format = "$#,##0"
    c_t = ws_dash.cell(row=r_offset, column=4, value=f_total)
    c_t.number_format = "$#,##0"
    c_p = ws_dash.cell(row=r_offset, column=5, value=f_ratio)
    c_p.number_format = "0.0%"
    
    for c in range(1, 6):
        ws_dash.cell(row=r_offset, column=c).border = thin_border

ws_adv = wb.create_sheet(title="Advanced Analytics")
ws_adv.merge_cells("A1:E1")
ws_adv["A1"] = "글로벌 시장 포트폴리오 4분면 수치 및 마진 구간"
ws_adv["A1"].font = font_title
ws_adv["A1"].fill = fill_navy
ws_adv["A1"].alignment = align_center

adv_headers = ["국가명 (Partner)", "총 수출액 ($)", "평균 단가 ($/kg)", "5개년 성장률 (%)", "포트폴리오 4분면 그룹"]
for c_idx, h in enumerate(adv_headers, start=1):
    cell = ws_adv.cell(row=3, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_center

adv_csv_path = 'BIZ-laver/docs/adv_stat_03_market_portfolio_matrix.csv'
if os.path.exists(adv_csv_path):
    df_adv = pd.read_csv(adv_csv_path)
    for r_idx, row in df_adv.iterrows():
        row_num = r_idx + 4
        ws_adv.cell(row=row_num, column=1, value=row['partnerDesc']).alignment = align_left
        
        c_v = ws_adv.cell(row=row_num, column=2, value=row['total_val'])
        c_v.number_format = "$#,##0"
        
        c_p = ws_adv.cell(row=row_num, column=3, value=row['avg_unit_price'])
        c_p.number_format = "$#,##0.00"
        
        c_g = ws_adv.cell(row=row_num, column=4, value=row['growth_rate'] / 100.0)
        c_g.number_format = "0.0%"
        
        ws_adv.cell(row=row_num, column=5, value=row['quadrant']).alignment = align_center
        
        for c in range(1, 6):
            ws_adv.cell(row=row_num, column=c).border = thin_border

for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(EXCEL_OUTPUT_PATH)
print(f"Excel 대시보드가 {EXCEL_OUTPUT_PATH}에 성공적으로 재생성 저장되었습니다.")
