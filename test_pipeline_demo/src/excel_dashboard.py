"""
test_pipeline_demo 엑셀 대시보드 구축 빌더

이 모듈은 openpyxl 라이브러리를 이용하여 크롤링된 CSV 데이터를 읽어,
깔끔하고 전문적인 디자인 서식, 피벗 성격의 요약 테이블, 
그리고 원본 데이터 시트가 갖추어진 엑셀 대시보드(.xlsx)를 생성합니다.

작성일: 2026-07-19
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def build_dashboard():
    print("[Excel Dashboard] 대시보드 구축을 시작합니다...")
    
    current_dir = os.path.dirname(os.path.abspath(__file__))
    data_path = os.path.abspath(os.path.join(current_dir, "..", "data", "raw_data.csv"))
    output_path = os.path.abspath(os.path.join(current_dir, "..", "data", "dashboard.xlsx"))
    
    if not os.path.exists(data_path):
        print(f"[Excel Dashboard] [ERROR] 원시 데이터가 존재하지 않습니다: {data_path}")
        return
        
    df = pd.read_csv(data_path)
    
    # openpyxl 워크북 생성
    wb = Workbook()
    
    # 1. 요약/대시보드 시트 생성
    ws_summary = wb.active
    ws_summary.title = "대시보드 요약"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # 스타일 정의
    title_font = Font(name="맑은 고딕", size=18, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    data_font = Font(name="맑은 고딕", size=10)
    thin_border = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )
    
    # 타이틀바 작성
    ws_summary.merge_cells("A1:E2")
    ws_summary["A1"] = "test_pipeline_demo 비즈니스 대시보드"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].fill = title_fill
    ws_summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    # 요약 지표 (KPI) 배치
    ws_summary["A4"] = "주요 지표"
    ws_summary["A4"].font = Font(name="맑은 고딕", size=13, bold=True)
    
    ws_summary["A5"] = "총 데이터 건수"
    ws_summary["B5"] = len(df)
    ws_summary["A5"].font = Font(name="맑은 고딕", size=11, bold=True)
    ws_summary["B5"].font = Font(name="맑은 고딕", size=11)
    
    # 2. 상세 데이터 시트 추가
    ws_data = wb.create_sheet(title="상세 데이터")
    ws_data.views.sheetView[0].showGridLines = True
    
    # 데이터프레임을 시트에 추가
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
        
    # 데이터 시트 스타일링 (헤더 적용)
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws_data.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # 열 너비 자동 맞춤
    for ws in [ws_summary, ws_data]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(output_path)
    print(f"[Excel Dashboard] 엑셀 대시보드 저장 완료: {output_path}")
    return output_path

if __name__ == "__main__":
    build_dashboard()
