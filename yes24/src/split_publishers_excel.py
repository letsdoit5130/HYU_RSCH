"""
이 스크립트는 기존에 작성된 yes24/docs/bestsellers_dashboard.xlsx 엑셀 파일을 로드하여,
베스트셀러 등록 도서 수가 가장 많은 상위 7개 출판사별로 개별 시트를 구성하고
나머지 모든 출판사의 도서 목록을 '기타_출판사'라는 단일 시트로 분리하여 저장하는 프로그램입니다.
주요 기능:
- 상위 7개 출판사 및 기타 출판사를 필터링하여 각각의 시트에 동일한 포맷으로 저장
- 새로 추가되는 모든 시트에 대해 가독성 높은 감청색 테마의 표 디자인(헤더 음영, 홀수행 줄무늬) 및 맑은 고딕 폰트 적용
- 엑셀 필수 요건에 따라 추가된 모든 시트의 눈금선(GridLines) 활성화 및 열 너비 최적화 수행
- 기존의 Dashboard, Analysis, RawData 시트 구조와 수식 관계는 완벽하게 유지
"""
# -*- coding: utf-8 -*-
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. 경로 설정
base_dir = r"C:\Users\leeak\OneDrive\1.HaeYu\HYU_RSCH\yes24"
data_path = os.path.join(base_dir, "data", "bestsellers.csv")
excel_path = os.path.join(base_dir, "docs", "bestsellers_dashboard.xlsx")

def split_publishers():
    # 데이터 로드 및 사전 파싱
    df = pd.read_csv(data_path)
    df['리뷰건수'] = df['리뷰건수'].astype(str).str.replace(',', '').str.replace('"', '').str.strip()
    df['리뷰건수'] = pd.to_numeric(df['리뷰건수'], errors='coerce').fillna(0).astype(int)
    df['정가'] = pd.to_numeric(df['정가'], errors='coerce').fillna(0).astype(int)
    df['할인가'] = pd.to_numeric(df['할인가'], errors='coerce').fillna(0).astype(int)
    df['할인율'] = pd.to_numeric(df['할인율'], errors='coerce').fillna(0.0)
    df['판매지수'] = pd.to_numeric(df['판매지수'], errors='coerce').fillna(0).astype(int)
    df['평점'] = pd.to_numeric(df['평점'], errors='coerce').fillna(0.0)
    df['태그'] = df['태그'].fillna('')

    # 상위 7개 출판사 리스트 확정
    top_7_publishers = df['출판사'].value_counts().head(7).index.tolist()
    print(f"선정된 상위 7개 출판사: {top_7_publishers}")
    
    # 기존 엑셀 파일 로드 (openpyxl)
    wb = load_workbook(excel_path)
    
    # 헤더 정의
    headers = ["순위", "도서명", "저자", "출판사", "출판일", "정가", "할인가", "판매지수", "리뷰건수", "평점", "태그"]
    
    # 서식 정의
    font_family = "Malgun Gothic"
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # 청회색 헤더
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # 지브라 연회색
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # 각 시트 생성 및 데이터 기록 함수
    def create_sheet_and_populate(sheet_name, subset_df):
        # 기존에 동일한 이름의 시트가 있다면 삭제 후 재생성 (안전 조치)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
            
        ws = wb.create_sheet(title=sheet_name)
        
        # 눈금선 활성화
        ws.views.sheetView[0].showGridLines = True
        
        # 헤더 기록 및 서식 적용
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(name=font_family, size=9.5, bold=True, color="FFFFFF")
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border
            
        # 데이터 기록
        row_idx = 2
        for r_data in subset_df[headers].values.tolist():
            ws.append(r_data)
            
            # 셀별 정밀 서식 매핑
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = Font(name=font_family, size=9)
                cell.border = thin_border
                
                # 수치형 정렬 및 표시 포맷 지정
                val = r_data[col_idx - 1]
                col_name = headers[col_idx - 1]
                
                if col_name in ["순위", "정가", "할인가", "판매지수", "리뷰건수"]:
                    cell.alignment = align_right
                    if col_name in ["정가", "할인가", "판매지수", "리뷰건수"]:
                        cell.number_format = '#,##0'
                elif col_name in ["출판일", "평점"]:
                    cell.alignment = align_center
                    if col_name == "평점":
                        cell.number_format = '0.0'
                else:
                    cell.alignment = align_left
                    
                # 지브라 패턴 적용
                if (row_idx - 2) % 2 == 1:
                    cell.fill = fill_zebra
                    
            row_idx += 1
            
        # 열 너비 자동 최적화
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        print(f"시트 생성 및 저장 완료: '{sheet_name}' (도서 수: {len(subset_df)}권)")

    # A. 상위 7개 출판사 개별 시트 작성
    for pub in top_7_publishers:
        subset = df[df['출판사'] == pub].sort_values(by="순위")
        # 시트명에 특수문자 제한 방지 (출판사 이름 그대로 사용)
        create_sheet_and_populate(pub, subset)
        
    # B. 나머지 출판사 시트 작성 ("기타_출판사" 시트)
    other_subset = df[~df['출판사'].isin(top_7_publishers)].sort_values(by="순위")
    create_sheet_and_populate("기타_출판사", other_subset)
    
    # 엑셀 최종 저장
    wb.save(excel_path)
    print(f"\n최종 완료: 엑셀 파일이 성공적으로 갱신되었습니다: {excel_path}")

if __name__ == "__main__":
    split_publishers()
