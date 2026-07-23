"""
이 스크립트는 yes24/data/bestsellers.csv 데이터를 로드하여
데이터 분석 결과와 시각화 요소를 포함한 다이내믹 Excel 대시보드를 생성하는 프로그램입니다.
주요 기능:
- RawData 시트: 베스트셀러 도서 목록의 모든 데이터 필드를 그대로 탑재
- Analysis 시트: 출판사별, 가격대별, 평점구간별 데이터를 COUNTIF/AVERAGEIFS 등 엑셀 수식으로 자동 집계
- Dashboard 시트: 감청색 테마의 세련된 디자인 요소를 적용하여 주요 KPI 카드(총 도서 수, 평균 판매지수 등)와 집계 데이터 매핑
- 대시보드 내에 openpyxl 기반의 실제 엑셀 막대 차트 2개(가격구간별 분포, TOP 10 출판사 실적)를 직접 생성 및 탑재
- 모든 시트의 눈금선(GridLines) 표시 설정 활성화 및 폰트 통일
"""
# -*- coding: utf-8 -*-
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# 1. 경로 설정
base_dir = r"C:\Users\leeak\OneDrive\1.HaeYu\HYU_RSCH\yes24"
data_path = os.path.join(base_dir, "data", "bestsellers.csv")
excel_output_path = os.path.join(base_dir, "docs", "bestsellers_dashboard.xlsx")

def build_dashboard():
    # 데이터 로드 및 사전 정제
    df = pd.read_csv(data_path)
    
    # 데이터 요약 정보 사전 추출 (출판사 정렬용)
    df['리뷰건수'] = df['리뷰건수'].astype(str).str.replace(',', '').str.replace('"', '').str.strip()
    df['리뷰건수'] = pd.to_numeric(df['리뷰건수'], errors='coerce').fillna(0).astype(int)
    df['정가'] = pd.to_numeric(df['정가'], errors='coerce').fillna(0).astype(int)
    df['판매지수'] = pd.to_numeric(df['판매지수'], errors='coerce').fillna(0).astype(int)
    df['평점'] = pd.to_numeric(df['평점'], errors='coerce').fillna(0.0)
    
    # 베스트셀러 빈도가 높은 출판사 순위 추출
    top_publishers = df['출판사'].value_counts().head(20).index.tolist()
    
    # openpyxl 워크북 생성
    wb = Workbook()
    
    # 기본 생성된 시트를 Dashboard로 지정
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    
    # 추가 시트 생성
    ws_anal = wb.create_sheet(title="Analysis")
    ws_raw = wb.create_sheet(title="RawData")
    
    # 눈금선 표시 설정 (스킬 필수 요건)
    for ws in [ws_dash, ws_anal, ws_raw]:
        ws.views.sheetView[0].showGridLines = True
        
    # --- 1단계: RawData 시트에 데이터 입력 ---
    # 헤더 입력
    headers = list(df.columns)
    ws_raw.append(headers)
    
    # 데이터 행 입력
    for r in df.values.tolist():
        # numeric 변환된 값 보정해서 로드
        row_data = []
        for idx, val in enumerate(r):
            col_name = headers[idx]
            if col_name == '리뷰건수':
                try:
                    row_data.append(int(str(val).replace(',', '')))
                except:
                    row_data.append(0)
            elif col_name in ['정가', '할인가', '할인율', '판매지수']:
                try:
                    row_data.append(int(float(val)))
                except:
                    row_data.append(0)
            elif col_name == '평점':
                try:
                    row_data.append(float(val))
                except:
                    row_data.append(0.0)
            else:
                row_data.append("" if pd.isna(val) else val)
        ws_raw.append(row_data)
        
    # --- 2단계: Analysis 시트에 요약 테이블 및 수식 세팅 ---
    # A. 출판사별 요약 (상위 20개 출판사)
    ws_anal['A1'] = "출판사"
    ws_anal['B1'] = "도서건수"
    ws_anal['C1'] = "평균판매지수"
    ws_anal['D1'] = "평균평점"
    ws_anal['E1'] = "총리뷰건수"
    
    for idx, pub in enumerate(top_publishers):
        row = idx + 2
        ws_anal[f'A{row}'] = pub
        ws_anal[f'B{row}'] = f"=COUNTIF(RawData!$F$2:$F$878, A{row})"
        ws_anal[f'C{row}'] = f"=AVERAGEIF(RawData!$F$2:$F$878, A{row}, RawData!$K$2:$K$878)"
        ws_anal[f'D{row}'] = f"=AVERAGEIF(RawData!$F$2:$F$878, A{row}, RawData!$M$2:$M$878)"
        ws_anal[f'E{row}'] = f"=SUMIF(RawData!$F$2:$F$878, A{row}, RawData!$L$2:$L$878)"
        
    # B. 가격대별 요약
    ws_anal['G1'] = "구간하한"
    ws_anal['H1'] = "구간상한"
    ws_anal['I1'] = "가격구간"
    ws_anal['J1'] = "도서건수"
    ws_anal['K1'] = "평균판매지수"
    
    price_ranges = [
        (0, 10000, "1만원 이하"),
        (10000, 15000, "1만~1.5만원"),
        (15000, 20000, "1.5만~2만원"),
        (20000, 25000, "2만~2.5만원"),
        (25000, 30000, "2.5만~3만원"),
        (30000, 999999, "3만원 초과")
    ]
    for idx, (low, high, label) in enumerate(price_ranges):
        row = idx + 2
        ws_anal[f'G{row}'] = low
        ws_anal[f'H{row}'] = high
        ws_anal[f'I{row}'] = label
        ws_anal[f'J{row}'] = f"=COUNTIFS(RawData!$H$2:$H$878, \">=\"&G{row}, RawData!$H$2:$H$878, \"<\"&H{row})"
        ws_anal[f'K{row}'] = f"=AVERAGEIFS(RawData!$K$2:$K$878, RawData!$H$2:$H$878, \">=\"&G{row}, RawData!$H$2:$H$878, \"<\"&H{row})"

    # C. 평점대별 요약
    ws_anal['M1'] = "평점하한"
    ws_anal['N1'] = "평점상한"
    ws_anal['O1'] = "평점구간"
    ws_anal['P1'] = "도서건수"
    ws_anal['Q1'] = "평균판매지수"
    
    rating_ranges = [
        (0, 8.0, "8점 이하"),
        (8.0, 9.0, "8점초과~9점"),
        (9.0, 9.5, "9점초과~9.5점"),
        (9.5, 10.1, "9.5점초과")
    ]
    for idx, (low, high, label) in enumerate(rating_ranges):
        row = idx + 2
        ws_anal[f'M{row}'] = low
        ws_anal[f'N{row}'] = high
        ws_anal[f'O{row}'] = label
        ws_anal[f'P{row}'] = f"=COUNTIFS(RawData!$M$2:$M$878, \">=\"&M{row}, RawData!$M$2:$M$878, \"<\"&N{row})"
        ws_anal[f'Q{row}'] = f"=AVERAGEIFS(RawData!$K$2:$K$878, RawData!$M$2:$M$878, \">=\"&M{row}, RawData!$M$2:$M$878, \"<\"&N{row})"


    # --- 3단계: Dashboard 시트 구축 및 UI 디자인 ---
    # 디자인 요소 설정
    font_family = "Malgun Gothic"
    
    # 테두리 선
    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    thick_bottom_side = Side(border_style="medium", color="2F5597")
    header_border = Border(bottom=thick_bottom_side, left=thin_border_side, right=thin_border_side, top=thin_border_side)
    
    # 색상 채우기
    fill_title = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # 감청색
    fill_kpi = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid") # 밝은 회푸른색
    fill_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # 청회색
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # 지브라 연회색
    
    # 정렬
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # A. 타이틀 바 생성 (A1:N2 병합)
    ws_dash.merge_cells("A1:K2")
    title_cell = ws_dash["A1"]
    title_cell.value = "  YES24 베스트셀러 탐색적 데이터 분석(EDA) & KPI 대시보드"
    title_cell.font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
    title_cell.fill = fill_title
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # B. 주요 KPI 카드 (A4 ~ K5 영역)
    kpis = [
        ("총 등록 도서 수", "=COUNTA(RawData!C2:C878)", "B4:C5", 'A4', '#,##0 "권"'),
        ("평균 판매지수", "=AVERAGE(RawData!K2:K878)", "E4:F5", 'D4', '#,##0'),
        ("평균 평점", "=AVERAGE(RawData!M2:M878)", "H4:I5", 'G4', '0.0 "점"'),
        ("총 리뷰 건수", "=SUM(RawData!L2:L878)", "K4:L5", 'J4', '#,##0 "건"')
    ]
    
    for label, formula, merge_range, label_cell_coord, num_fmt in kpis:
        # 셀 병합 및 스타일링
        ws_dash.merge_cells(merge_range)
        val_cell = ws_dash[merge_range.split(":")[0]]
        val_cell.value = formula
        val_cell.font = Font(name=font_family, size=13, bold=True, color="1F4E79")
        val_cell.alignment = align_center
        val_cell.number_format = num_fmt
        
        # 라벨 셀 생성 (값 셀 바로 위)
        lbl_cell = ws_dash[label_cell_coord]
        lbl_cell.value = label
        lbl_cell.font = Font(name=font_family, size=9, bold=True, color="595959")
        lbl_cell.alignment = align_center
        
        # 카드 영역 배경색 및 테두리 (병합 영역의 각 셀에 설정)
        start_col, start_row = merge_range.split(":")[0][0], int(merge_range.split(":")[0][1:])
        end_col, end_row = merge_range.split(":")[1][0], int(merge_range.split(":")[1][1:])
        
        # 라벨 및 값 영역 전체에 스타일 입히기
        col_start_idx = ord(start_col) - ord('A') + 1
        col_end_idx = ord(end_col) - ord('A') + 1
        
        for r_idx in range(start_row - 1, end_row + 1):
            for c_idx in range(col_start_idx, col_end_idx + 1):
                cell = ws_dash.cell(row=r_idx, column=c_idx)
                cell.fill = fill_kpi
                cell.border = thin_border

    # C. 대시보드 테이블 1: TOP 10 출판사 실적 요약
    ws_dash['A7'] = "TOP 10 출판사 요약"
    ws_dash['A7'].font = Font(name=font_family, size=11, bold=True, color="1F4E79")
    
    headers_t1 = ["출판사명", "도서 건수", "평균 판매지수", "평균 평점", "총 리뷰건수"]
    for c_idx, h in enumerate(headers_t1):
        cell = ws_dash.cell(row=8, column=c_idx+1)
        cell.value = h
        cell.font = Font(name=font_family, size=9.5, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = header_border
        
    for idx in range(10):
        r_idx = idx + 9
        anal_row = idx + 2
        # 수식 매핑 (동적 연동)
        ws_dash.cell(row=r_idx, column=1, value=f"=Analysis!A{anal_row}").alignment = align_left
        ws_dash.cell(row=r_idx, column=2, value=f"=Analysis!B{anal_row}").number_format = '#,##0'
        ws_dash.cell(row=r_idx, column=3, value=f"=Analysis!C{anal_row}").number_format = '#,##0'
        ws_dash.cell(row=r_idx, column=4, value=f"=Analysis!D{anal_row}").number_format = '0.0'
        ws_dash.cell(row=r_idx, column=5, value=f"=Analysis!E{anal_row}").number_format = '#,##0'
        
        for col in range(1, 6):
            cell = ws_dash.cell(row=r_idx, column=col)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            if col > 1 and col != 4:
                cell.alignment = align_right
            elif col == 4:
                cell.alignment = align_center
            if idx % 2 == 1:
                cell.fill = fill_zebra

    # D. 대시보드 테이블 2: 도서 가격대별 분포
    ws_dash['G7'] = "도서 가격대별 분포"
    ws_dash['G7'].font = Font(name=font_family, size=11, bold=True, color="1F4E79")
    
    headers_t2 = ["가격구간", "도서 건수", "평균 판매지수"]
    for c_idx, h in enumerate(headers_t2):
        cell = ws_dash.cell(row=8, column=c_idx+7)
        cell.value = h
        cell.font = Font(name=font_family, size=9.5, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = header_border
        
    for idx in range(6):
        r_idx = idx + 9
        anal_row = idx + 2
        ws_dash.cell(row=r_idx, column=7, value=f"=Analysis!I{anal_row}").alignment = align_center
        ws_dash.cell(row=r_idx, column=8, value=f"=Analysis!J{anal_row}").number_format = '#,##0'
        ws_dash.cell(row=r_idx, column=9, value=f"=Analysis!K{anal_row}").number_format = '#,##0'
        
        for col in range(7, 10):
            cell = ws_dash.cell(row=r_idx, column=col)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            if col > 7:
                cell.alignment = align_right
            if idx % 2 == 1:
                cell.fill = fill_zebra

    # E. 대시보드 테이블 3: 평점대별 분포
    ws_dash['G16'] = "도서 평점대별 분포"
    ws_dash['G16'].font = Font(name=font_family, size=11, bold=True, color="1F4E79")
    
    headers_t3 = ["평점구간", "도서 건수", "평균 판매지수"]
    for c_idx, h in enumerate(headers_t3):
        cell = ws_dash.cell(row=17, column=c_idx+7)
        cell.value = h
        cell.font = Font(name=font_family, size=9.5, bold=True, color="FFFFFF")
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = header_border
        
    for idx in range(4):
        r_idx = idx + 18
        anal_row = idx + 2
        ws_dash.cell(row=r_idx, column=7, value=f"=Analysis!O{anal_row}").alignment = align_center
        ws_dash.cell(row=r_idx, column=8, value=f"=Analysis!P{anal_row}").number_format = '#,##0'
        ws_dash.cell(row=r_idx, column=9, value=f"=Analysis!Q{anal_row}").number_format = '#,##0'
        
        for col in range(7, 10):
            cell = ws_dash.cell(row=r_idx, column=col)
            cell.font = Font(name=font_family, size=9)
            cell.border = thin_border
            if col > 7:
                cell.alignment = align_right
            if idx % 2 == 1:
                cell.fill = fill_zebra

    # --- 4단계: openpyxl Chart 추가 (Dashboard 시트에 엑셀 차트 삽입) ---
    # 차트 1: 가격구간별 도서 건수 세로 막대 그래프 (Column Chart)
    chart_price = BarChart()
    chart_price.type = "col"
    chart_price.title = "가격대별 베스트셀러 도서 건수 분포"
    chart_price.style = 10
    chart_price.y_axis.title = "도서 건수 (권)"
    chart_price.x_axis.title = "가격구간"
    chart_price.width = 16
    chart_price.height = 8.5
    
    data_price = Reference(ws_anal, min_col=10, min_row=1, max_row=7) # J열 (도서건수)
    cats_price = Reference(ws_anal, min_col=9, min_row=2, max_row=7)  # I열 (가격대)
    chart_price.add_data(data_price, titles_from_data=True)
    chart_price.set_categories(cats_price)
    chart_price.legend = None # 단일 계열이므로 범례 생략
    ws_dash.add_chart(chart_price, "M8")
    
    # 차트 2: TOP 10 출판사별 평균 판매지수 가로 막대 그래프 (Bar Chart)
    chart_pub = BarChart()
    chart_pub.type = "bar"
    chart_pub.title = "TOP 10 출판사별 평균 판매지수"
    chart_pub.style = 13
    chart_pub.x_axis.title = "출판사명"
    chart_pub.y_axis.title = "평균 판매지수"
    chart_pub.width = 16
    chart_pub.height = 9.5
    
    data_pub = Reference(ws_anal, min_col=3, min_row=1, max_row=11) # C열 (평균판매지수)
    cats_pub = Reference(ws_anal, min_col=1, min_row=2, max_row=11)  # A열 (출판사)
    chart_pub.add_data(data_pub, titles_from_data=True)
    chart_pub.set_categories(cats_pub)
    chart_pub.legend = None
    ws_dash.add_chart(chart_pub, "M26")

    # --- 5단계: 열 너비 최적화 ---
    # Dashboard 열 너비 명시적 할당
    col_widths_dash = {
        'A': 18, 'B': 12, 'C': 15, 'D': 12, 'E': 14, 
        'F': 12, 'G': 18, 'H': 12, 'I': 15, 'J': 12, 'K': 14
    }
    for col_let, w in col_widths_dash.items():
        ws_dash.column_dimensions[col_let].width = w
        
    # Analysis 열 너비 명시적 할당
    col_widths_anal = {
        'A': 20, 'B': 12, 'C': 15, 'D': 12, 'E': 12, 
        'G': 12, 'H': 12, 'I': 15, 'J': 12, 'K': 15,
        'M': 12, 'N': 12, 'O': 18, 'P': 12, 'Q': 15
    }
    for col_let, w in col_widths_anal.items():
        ws_anal.column_dimensions[col_let].width = w
        
    # RawData 자동 열 너비 세팅
    for col in ws_raw.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_raw.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # 워크북 최종 저장
    wb.save(excel_output_path)
    print(f"성공: Excel 대시보드가 작성되어 다음 경로에 저장되었습니다: {excel_output_path}")

if __name__ == "__main__":
    build_dashboard()
