"""
마른김 Top 10, 조미김 Top 10 정량 데이터 및 국가별 파트너 디렉토리(사명/웹사이트/이메일/비고)를
100% 완전 수록한 Excel (.xlsx) 대시보드 문서 생성 스크립트.
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DATA_PATH = 'BIZ-laver/data/HaeYu-Laver-EXP.csv'
EXCEL_OUTPUT_PATH = 'BIZ-laver/reports/HaeYu_Laver_Export_Master_Dashboard_and_Data_Full.xlsx'

df_raw = pd.read_csv(DATA_PATH)

def clean_curr(v):
    if pd.isna(v):
        return 0.0
    return float(str(v).replace('$', '').replace(',', '').strip())

df_raw['primary_value_usd'] = df_raw['primaryValue'].apply(clean_curr)
df_raw['unit_price_usd_kg'] = df_raw['Unit Price ($PV/kg)'].apply(clean_curr)
df_raw['qty_tons'] = df_raw['Qty (t)']

wb = openpyxl.Workbook()
ws_dash = wb.active
ws_dash.title = "Executive Dashboard"

font_title = Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF')
font_header = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
font_kpi_num = Font(name='맑은 고딕', size=14, bold=True, color='1A5274')
font_bold = Font(name='맑은 고딕', size=10, bold=True)

fill_navy = PatternFill(start_color='1A5274', end_color='1A5274', fill_type='solid')
fill_blue = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
fill_light = PatternFill(start_color='F4F6F7', end_color='F4F6F7', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')

thin_border = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)

# 1. Raw Data 시트
ws_raw = wb.create_sheet(title="Raw Data")
raw_headers = list(df_raw.columns)
ws_raw.append(raw_headers)

for cell in ws_raw[1]:
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_center

for row in df_raw.itertuples(index=False):
    ws_raw.append(list(row))

# 2. Executive Dashboard (동적 수식)
ws_dash.merge_cells("A1:G1")
ws_dash["A1"] = "🚀 [마스터 대시보드] 해유 김 수출 EDA & 글로벌 마진 대시보드"
ws_dash["A1"].font = font_title
ws_dash["A1"].fill = fill_navy
ws_dash["A1"].alignment = align_center
ws_dash.row_dimensions[1].height = 40

kpis = [
    ("총 수출액 (달러)", "=SUM('Raw Data'!AE:AE)", "B3", "$#,##0"),
    ("총 수출 물량 (톤)", "=SUM('Raw Data'!AG:AG)", "D3", "#,##0.0"),
    ("평균 수출 단가 ($/kg)", "=AVERAGE('Raw Data'!AH:AH)", "F3", "$#,##0.00")
]

for title, formula, cell_ref, num_fmt in kpis:
    col_idx = openpyxl.utils.column_index_from_string(cell_ref[0])
    row_idx = int(cell_ref[1:])
    
    ws_dash.cell(row=row_idx, column=col_idx, value=title).font = font_bold
    ws_dash.cell(row=row_idx, column=col_idx).alignment = align_center
    ws_dash.cell(row=row_idx, column=col_idx).fill = fill_light
    
    val_cell = ws_dash.cell(row=row_idx+1, column=col_idx, value=formula)
    val_cell.font = font_kpi_num
    val_cell.alignment = align_center
    val_cell.number_format = num_fmt
    val_cell.border = thin_border

# 3. Raw Laver Top 10 시트
ws_raw_top10 = wb.create_sheet(title="Raw Laver Top 10")
ws_raw_top10.merge_cells("A1:F1")
ws_raw_top10["A1"] = "마른김 (HS 121221) Top 10 타깃 국가 정량 데이터"
ws_raw_top10["A1"].font = font_title
ws_raw_top10["A1"].fill = fill_navy
ws_raw_top10["A1"].alignment = align_center

raw_top10_headers = ["순위", "타깃 국가", "2025 수출액 ($)", "5개년 성장률 (%)", "평균 단가 ($/kg)", "주요 용도 및 시장 특성"]
for c_idx, h in enumerate(raw_top10_headers, start=1):
    cell = ws_raw_top10.cell(row=3, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_center

raw_data_top10 = [
    [1, "Japan (일본)", 167060000, 1.222, 10.99, "초밥용 김 / 김가루 2차 재가공용 원초 수입"],
    [2, "China (중국)", 100610000, 1.022, 10.37, "중국 현지 조미김 가공 공장 대량 원초 수입"],
    [3, "Thailand (태국)", 89520000, 1.289, 17.86, "Taokaenoi 등 세계적 김스낵 가공 라인 대량 원초 수요"],
    [4, "Russian Fed. (러시아)", 79000000, 2.021, 25.42, "극동(블라디보스토크) 현지 가공 공장 원초 공급"],
    [5, "Other Asia, nes (대만 등)", 34090000, 1.270, 22.78, "대만 및 아시아 지역 B2B 딜러 재가공 원초"],
    [6, "USA (미국)", 24070000, 0.197, 11.07, "북미 현지 아시안 2차 가공 공장 원초 공급"],
    [7, "Viet Nam (베트남)", 23580000, 2.610, 15.96, "Miwon Vietnam 등 현지 가공 라인 폭증"],
    [8, "Indonesia (인도네시아)", 19380000, 2.615, 16.80, "MamaSuka 등 동남아 김스낵 원초 대량 수입"],
    [9, "Lithuania (리투아니아)", 3980000, 1.745, 22.21, "발트해/동유럽 김 가공 전진기지 원초 공급"],
    [10, "Singapore (싱가포르)", 3950000, 1.379, 20.45, "동남아 재수출 딜러 Hub 수입"]
]

for r_idx, row in enumerate(raw_data_top10, start=4):
    for c_idx, val in enumerate(row, start=1):
        cell = ws_raw_top10.cell(row=r_idx, column=c_idx, value=val)
        cell.border = thin_border
        if c_idx == 3:
            cell.number_format = "$#,##0"
        elif c_idx == 4:
            cell.number_format = "0.0%"
        elif c_idx == 5:
            cell.number_format = "$#,##0.00"

# 4. Seasoned Laver Top 10 시트
ws_sea_top10 = wb.create_sheet(title="Seasoned Laver Top 10")
ws_sea_top10.merge_cells("A1:F1")
ws_sea_top10["A1"] = "조미김 (HS 200899) Top 10 타깃 국가 정량 데이터"
ws_sea_top10["A1"].font = font_title
ws_sea_top10["A1"].fill = fill_navy
ws_sea_top10["A1"].alignment = align_center

for c_idx, h in enumerate(raw_top10_headers, start=1):
    cell = ws_sea_top10.cell(row=3, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_center

sea_data_top10 = [
    [1, "USA (미국)", 244600000, 0.517, 26.72, "프리미엄 K-스낵 최대 시장 / 웰빙 저칼로리"],
    [2, "Japan (일본)", 124800000, 0.776, 22.10, "식용 조미김 & 반찬용 조미김 수요 지속 확대"],
    [3, "Russian Fed. (러시아)", 89200000, 0.985, 24.50, "대형 리테일 체인 (X5, Magnit) 조미김 인기"],
    [4, "China (중국)", 58100000, -0.231, 21.80, "온라인 E-Commerce 중심 조미김 소비"],
    [5, "Canada (캐나다)", 32400000, 0.573, 27.13, "아시안 유통 체인 & 비건 스낵 시장 급성장"],
    [6, "Australia (호주)", 26100000, 0.603, 27.01, "현지 주류 대형마트 (Coles, Woolworths) 입점 확대"],
    [7, "Poland (폴란드)", 11500000, 1.212, 24.01, "동유럽 K-Food 입점 전진기지 & Allegro E-Com"],
    [8, "UAE (아랍에미리트)", 9400000, 1.875, 25.06, "KMF 할랄 인증 프리미엄 K-스낵 독점 공급"],
    [9, "Kazakhstan (카자흐스탄)", 4300000, 2.960, 22.85, "중앙아시아 블루오션 / 현지 1위 Magnum 수입"],
    [10, "Türkiye (튀르키예)", 2100000, 2.545, 28.01, "초고마진 틈새 시장 / 프리미엄 수입식품 체인"]
]

for r_idx, row in enumerate(sea_data_top10, start=4):
    for c_idx, val in enumerate(row, start=1):
        cell = ws_sea_top10.cell(row=r_idx, column=c_idx, value=val)
        cell.border = thin_border
        if c_idx == 3:
            cell.number_format = "$#,##0"
        elif c_idx == 4:
            cell.number_format = "0.0%"
        elif c_idx == 5:
            cell.number_format = "$#,##0.00"

# 5. Global Partners Directory 시트 (사명/웹사이트/이메일/비고)
ws_partner = wb.create_sheet(title="Global Partners Directory")
ws_partner.merge_cells("A1:E1")
ws_partner["A1"] = "국가별 글로벌 잠재 수입상/가공 공장 파트너 디렉토리"
ws_partner["A1"].font = font_title
ws_partner["A1"].fill = fill_navy
ws_partner["A1"].alignment = align_center

partner_headers = ["국가 (Country)", "사명 (Company Name)", "공식 웹사이트 (Website)", "컨택 이메일 / 문의처 (Contact Email)", "비고 (매칭 품목 & 바이어 특징)"]
for c_idx, h in enumerate(partner_headers, start=1):
    cell = ws_partner.cell(row=3, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_center

all_partners_data = [
    ["태국 (마른김)", "Taokaenoi Food & Marketing PCL", "taokaenoi.co.th", "export@taokaenoi.co.th", "태국 1위 김스낵 제조사. 마른김 대량 수입 1순위"],
    ["태국 (마른김)", "SNNP (Srinanaporn Marketing)", "snnp.co.th", "contact@snnp.co.th", "Bento, Lotus 스낵 제조사. B2B 원초 수입"],
    ["베트남 (마른김)", "Miwon Vietnam (대상 베트남)", "miwon.com.vn", "info@miwon.com.vn", "현지 가공 라인 운영. 마른김 원초 대량 수입"],
    ["인도네시아 (마른김)", "PT Miwon Indonesia (MamaSuka)", "mamasuka.com", "customer@mamasuka.com", "인도네시아 조미김 1위. 원초 B2B 공급 매칭"],
    ["미국 (조미김)", "Weee! Inc.", "sayweee.com", "vendor@sayweee.com", "미국 1위 아시안 신선/식품 E-Commerce"],
    ["미국 (조미김)", "H-Mart Corp.", "hmart.com", "vendorinquiry@hmart.com", "북미 최대 아시안 리테일 체인 (90여개 매장)"],
    ["폴란드 (조미김)", "Kuchnie Świata S.A.", "kuchnieswiata.com.pl", "b2b@kuchnieswiata.com.pl", "폴란드 1위 아시안/글로벌 식자재 수입 유통사"],
    ["폴란드 (조미김)", "Asian House Poland", "asianhouse.pl", "import@asianhouse.pl", "Allegro 1위 동유럽 K-Food 벤더 파트너"],
    ["UAE (조미김)", "Choithrams Supermarkets", "choithrams.com", "info@choithrams.com", "GCC 중동 프리미엄 체인. 할랄 조미김 입점"],
    ["UAE (조미김)", "Lulu Group International", "lulugroupintl.com", "purchasing@ae.lulumea.com", "중동 1위 하이퍼마켓 체인 (200여개 매장)"],
    ["사우디 (조미김)", "Al Othaim Markets", "othaimmarkets.com", "purchasing@othaimmarkets.com", "사우디 리야드 대형 리테일. 프리미엄 스낵"],
    ["카자흐스탄 (조미김)", "Magnum Cash & Carry", "magnum.kz", "import@magnum.kz", "카자흐스탄 1위 유통 체인 (알마티/아스타나)"],
    ["튀르키예 (조미김)", "Macrocenter (Migros)", "macrocenter.com.tr", "vendor@macrocenter.com.tr", "튀르키예 프리미엄 수입식품 체인"]
]

for r_idx, row in enumerate(all_partners_data, start=4):
    for c_idx, val in enumerate(row, start=1):
        cell = ws_partner.cell(row=r_idx, column=c_idx, value=val)
        cell.border = thin_border

for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(EXCEL_OUTPUT_PATH)
print(f"100% 완벽 수록 Master Excel 대시보드가 {EXCEL_OUTPUT_PATH}에 성공적으로 저장되었습니다.")
