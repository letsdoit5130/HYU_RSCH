"""
교보문고 베스트셀러 엑셀 대시보드 빌더

이 모듈은 KyoBooks/data/bestsellers.csv 데이터를 로드하여
다양한 비즈니스 지표 요약 및 시각화 요소를 포함한 Excel 대시보드를 구축합니다.
최종 산출물은 KyoBooks/docs/bestsellers_dashboard.xlsx로 저장됩니다.

주요 기능:
1. RawData 시트: 모든 수집 데이터 항목 기록 및 포맷팅 (정수/실수형 변환 포함)
2. Analysis 시트: 출판사별, 가격구간별, 평점구간별 데이터를 COUNTIF/AVERAGEIFS 등 엑셀 수식으로 집계
3. Dashboard 시트: 교보문고 브랜드 상징색(Dark Green: #004F2F) 중심의 UI 스타일 적용,
   주요 KPI 카드 4종 구성, 집계 테이블 매핑, 그리고 오픈파이엑셀(openpyxl) 내장 막대 차트 2종(출판사별 점유율, 가격구간별 분포) 배치
4. 모든 시트 눈금선(GridLines) 표시 활성화 및 열 너비 자동 조정
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

def build_dashboard():
    # 경로 설정
    current_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(current_dir)
    data_path = os.path.join(project_dir, "data", "bestsellers.csv")
    excel_output_path = os.path.join(project_dir, "docs", "bestsellers_dashboard.xlsx")
    os.makedirs(os.path.dirname(excel_output_path), exist_ok=True)
    
    if not os.path.exists(data_path):
        print(f"데이터 파일이 존재하지 않습니다: {data_path}")
        return
        
    df = pd.read_csv(data_path, encoding="utf-8-sig")
    total_rows = len(df)
    
    # 데이터 사전 전처리
    df['정가'] = pd.to_numeric(df['정가'], errors='coerce').fillna(0).astype(int)
    df['할인가'] = pd.to_numeric(df['할인가'], errors='coerce').fillna(0).astype(int)
    df['할인율'] = pd.to_numeric(df['할인율'], errors='coerce').fillna(0).astype(int)
    df['리뷰건수'] = pd.to_numeric(df['리뷰건수'], errors='coerce').fillna(0).astype(int)
    df['평점'] = pd.to_numeric(df['평점'], errors='coerce').fillna(0.0)
    
    # 상위 출판사 리스트 추출
    top_publishers = df['출판사'].value_counts().head(15).index.tolist()
    
    # 엑셀 워크북 생성
    wb = Workbook()
    
    # 시트 구성
    ws_dash = wb.active
    ws_dash.title = "Dashboard"
    ws_anal = wb.create_sheet(title="Analysis")
    ws_raw = wb.create_sheet(title="RawData")
    
    # 눈금선 표시 활성화
    for ws in [ws_dash, ws_anal, ws_raw]:
        ws.views.sheetView[0].showGridLines = True
        
    # -------------------------------------------------------------
    # 1. RawData 시트 작성
    # -------------------------------------------------------------
    headers = list(df.columns)
    ws_raw.append(headers)
    
    for row in df.values.tolist():
        # 수치형 정제된 데이터 반영
        row_data = []
        for idx, val in enumerate(row):
            col_name = headers[idx]
            if col_name in ['정가', '할인가', '할인율', '판매지수', '리뷰건수']:
                try:
                    row_data.append(int(float(val)))
                except:
                    row_data.append(0)
            elif col_name == '평점':
                try:
                    row_data.append(float(val))
                except:
                    row_data.append(0.0)
            else:
                row_data.append("" if pd.isna(val) else val)
        ws_raw.append(row_data)
        
    # 헤더 스타일링 (Dark Green 테마)
    header_fill = PatternFill(start_color="004F2F", end_color="004F2F", fill_type="solid")
    header_font = Font(name="Malgun Gothic", size=10, bold=True, color="FFFFFF")
    for col_idx in range(1, len(headers) + 1):
        cell = ws_raw.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # -------------------------------------------------------------
    # 2. Analysis 시트 작성 (수식 집계)
    # -------------------------------------------------------------
    # A. 출판사별 요약
    ws_anal['A1'] = "출판사"
    ws_anal['B1'] = "도서건수"
    ws_anal['C1'] = "평균평점"
    ws_anal['D1'] = "총리뷰건수"
    
    last_row_raw = total_rows + 1
    
    for idx, pub in enumerate(top_publishers):
        row = idx + 2
        ws_anal[f'A{row}'] = pub
        ws_anal[f'B{row}'] = f"=COUNTIF(RawData!$F$2:$F${last_row_raw}, A{row})"
        ws_anal[f'C{row}'] = f"=AVERAGEIF(RawData!$F$2:$F${last_row_raw}, A{row}, RawData!$M$2:$M${last_row_raw})"
        ws_anal[f'D{row}'] = f"=SUMIF(RawData!$F$2:$F${last_row_raw}, A{row}, RawData!$L$2:$L${last_row_raw})"
        
    # B. 가격대별 요약
    ws_anal['F1'] = "구간하한"
    ws_anal['G1'] = "구간상한"
    ws_anal['H1'] = "가격구간"
    ws_anal['I1'] = "도서건수"
    ws_anal['J1'] = "평균평점"
    
    price_ranges = [
        (0, 10000, "1만원 이하"),
        (10000, 15000, "1만~1.5만원"),
        (15000, 20000, "1.5만~2만원"),
        (20000, 25000, "2만~2.5만원"),
        (25000, 30000, "2.5만~3만원"),
        (30000, 999999, "3만원 초과")
    ]
    for idx, (low, high, label) in enumerate(price_ranges):
        row = idx + 2
        ws_anal[f'F{row}'] = low
        ws_anal[f'G{row}'] = high
        ws_anal[f'H{row}'] = label
        ws_anal[f'I{row}'] = f"=COUNTIFS(RawData!$I$2:$I${last_row_raw}, \">=\"&F{row}, RawData!$I$2:$I${last_row_raw}, \"<\"&G{row})"
        ws_anal[f'J{row}'] = f"=AVERAGEIFS(RawData!$M$2:$M${last_row_raw}, RawData!$I$2:$I${last_row_raw}, \">=\"&F{row}, RawData!$I$2:$I${last_row_raw}, \"<\"&G{row})"
        
    # C. 평점대별 요약
    ws_anal['L1'] = "평점하한"
    ws_anal['M1'] = "평점상한"
    ws_anal['N1'] = "평점구간"
    ws_anal['O1'] = "도서건수"
    ws_anal['P1'] = "평균리뷰수"
    
    rating_ranges = [
        (0.0, 8.0, "8.0 이하"),
        (8.0, 9.0, "8.0 초과~9.0"),
        (9.0, 9.5, "9.0 초과~9.5"),
        (9.5, 10.1, "9.5 초과")
    ]
    for idx, (low, high, label) in enumerate(rating_ranges):
        row = idx + 2
        ws_anal[f'L{row}'] = low
        ws_anal[f'M{row}'] = high
        ws_anal[f'N{row}'] = label
        ws_anal[f'O{row}'] = f"=COUNTIFS(RawData!$M$2:$M${last_row_raw}, \">=\"&L{row}, RawData!$M$2:$M${last_row_raw}, \"<\"&M{row})"
        ws_anal[f'P{row}'] = f"=AVERAGEIFS(RawData!$L$2:$L${last_row_raw}, RawData!$M$2:$M${last_row_raw}, \">=\"&L{row}, RawData!$M$2:$M${last_row_raw}, \"<\"&M{row})"

    # Analysis 시트 헤더 스타일링
    anal_header_fill = PatternFill(start_color="336633", end_color="336633", fill_type="solid")
    for col_idx in [1,2,3,4, 6,7,8,9,10, 12,13,14,15,16]:
        cell = ws_anal.cell(row=1, column=col_idx)
        cell.fill = anal_header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # -------------------------------------------------------------
    # 3. Dashboard 시트 디자인 및 배치
    # -------------------------------------------------------------
    font_name = "Malgun Gothic"
    
    # 스타일 선언
    title_font = Font(name=font_name, size=18, bold=True, color="004F2F")
    section_font = Font(name=font_name, size=12, bold=True, color="333333")
    kpi_title_font = Font(name=font_name, size=9, color="666666")
    kpi_value_font = Font(name=font_name, size=16, bold=True, color="004F2F")
    
    border_side = Side(border_style="thin", color="D9D9D9")
    box_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    fill_kpi = PatternFill(start_color="F2F7F4", end_color="F2F7F4", fill_type="solid")
    fill_header = PatternFill(start_color="004F2F", end_color="004F2F", fill_type="solid")
    
    # A. 타이틀
    ws_dash.merge_cells("A1:N2")
    title_cell = ws_dash["A1"]
    title_cell.value = "교보문고 실시간 베스트셀러 분석 대시보드"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # B. KPI 카드 배치 (3행~4행)
    kpis = [
        ("총 도서 수", f"=COUNTA(RawData!C2:C{last_row_raw})", "A3", "B3", "A4", "B4"),
        ("평균 정가", f"=AVERAGE(RawData!H2:H{last_row_raw})", "D3", "E3", "D4", "E4"),
        ("평균 평점", f"=AVERAGE(RawData!M2:M{last_row_raw})", "G3", "H3", "G4", "H4"),
        ("총 리뷰 수", f"=SUM(RawData!L2:L{last_row_raw})", "J3", "K3", "J4", "K4")
    ]
    
    for label, formula, t_start, t_end, v_start, v_end in kpis:
        # 타이틀 병합 및 기입
        ws_dash.merge_cells(f"{t_start}:{t_end}")
        t_cell = ws_dash[t_start]
        t_cell.value = label
        t_cell.font = kpi_title_font
        t_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 값 병합 및 기입
        ws_dash.merge_cells(f"{v_start}:{v_end}")
        v_cell = ws_dash[v_start]
        v_cell.value = formula
        v_cell.font = kpi_value_font
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 테두리 및 배경색 적용
        col_letter = t_start[0]
        col_idx = ord(col_letter) - ord('A') + 1
        for r in range(3, 5):
            for c in range(col_idx, col_idx + 2):
                cell = ws_dash.cell(row=r, column=c)
                cell.border = box_border
                cell.fill = fill_kpi
                
    # C. 데이터 요약 테이블 배치 (6행부터)
    # 출판사 TOP 10 요약 테이블
    ws_dash["A6"] = "점유율 상위 출판사 TOP 10"
    ws_dash["A6"].font = section_font
    
    ws_dash["A7"] = "출판사"
    ws_dash["B7"] = "도서건수"
    ws_dash["C7"] = "평균평점"
    ws_dash["D7"] = "총리뷰건수"
    
    for col in ["A", "B", "C", "D"]:
        ws_dash[f"{col}7"].fill = fill_header
        ws_dash[f"{col}7"].font = header_font
        ws_dash[f"{col}7"].alignment = Alignment(horizontal="center")
        
    for i in range(10):
        row = i + 8
        anal_row = i + 2
        ws_dash[f"A{row}"] = f"=Analysis!A{anal_row}"
        ws_dash[f"B{row}"] = f"=Analysis!B{anal_row}"
        ws_dash[f"C{row}"] = f"=Analysis!C{anal_row}"
        ws_dash[f"D{row}"] = f"=Analysis!D{anal_row}"
        
        # 서식 및 테두리 적용
        ws_dash[f"C{row}"].number_format = '0.0'
        ws_dash[f"D{row}"].number_format = '#,##0'
        for col in ["A", "B", "C", "D"]:
            ws_dash[f"{col}{row}"].border = box_border
            ws_dash[f"{col}{row}"].font = Font(name=font_name, size=9)
            ws_dash[f"{col}{row}"].alignment = Alignment(horizontal="center" if col != "A" else "left")
            
    # 가격대별 요약 테이블
    ws_dash["F6"] = "가격대별 분포 현황"
    ws_dash["F6"].font = section_font
    
    ws_dash["F7"] = "가격구간"
    ws_dash["G7"] = "도서건수"
    ws_dash["H7"] = "평균평점"
    
    for col in ["F", "G", "H"]:
        ws_dash[f"{col}7"].fill = fill_header
        ws_dash[f"{col}7"].font = header_font
        ws_dash[f"{col}7"].alignment = Alignment(horizontal="center")
        
    for i in range(6):
        row = i + 8
        anal_row = i + 2
        ws_dash[f"F{row}"] = f"=Analysis!H{anal_row}"
        ws_dash[f"G{row}"] = f"=Analysis!I{anal_row}"
        ws_dash[f"H{row}"] = f"=Analysis!J{anal_row}"
        
        # 서식 및 테두리 적용
        ws_dash[f"H{row}"].number_format = '0.0'
        for col in ["F", "G", "H"]:
            ws_dash[f"{col}{row}"].border = box_border
            ws_dash[f"{col}{row}"].font = Font(name=font_name, size=9)
            ws_dash[f"{col}{row}"].alignment = Alignment(horizontal="center")

    # D. openpyxl 차트 생성 및 배치
    # 차트 1: 출판사 TOP 10 실적
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "출판사 TOP 10 베스트셀러 수"
    chart1.y_axis.title = "도서 수 (권)"
    chart1.x_axis.title = "출판사"
    chart1.legend = None  # 범례 불필요
    chart1.width = 16
    chart1.height = 10
    
    data_ref1 = Reference(ws_dash, min_col=2, min_row=7, max_row=17) # 도서건수
    cats_ref1 = Reference(ws_dash, min_col=1, min_row=8, max_row=17) # 출판사명
    chart1.add_data(data_ref1, titles_from_data=True)
    chart1.set_categories(cats_ref1)
    ws_dash.add_chart(chart1, "A19")
    
    # 차트 2: 가격구간별 분포
    chart2 = BarChart()
    chart2.type = "col"
    chart2.style = 11
    chart2.title = "가격구간별 도서 수 분포"
    chart2.y_axis.title = "도서 수 (권)"
    chart2.x_axis.title = "가격대"
    chart2.legend = None
    chart2.width = 14
    chart2.height = 10
    
    data_ref2 = Reference(ws_dash, min_col=7, min_row=7, max_row=13) # 도서건수
    cats_ref2 = Reference(ws_dash, min_col=6, min_row=8, max_row=13) # 가격구간명
    chart2.add_data(data_ref2, titles_from_data=True)
    chart2.set_categories(cats_ref2)
    ws_dash.add_chart(chart2, "F19")
    
    # -------------------------------------------------------------
    # 4. 열 너비 자동 조정 공통 적용
    # -------------------------------------------------------------
    for ws in [ws_dash, ws_anal, ws_raw]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # 셀 수식이나 문자열 길이를 바탕으로 적당한 너비 계산
            for cell in col:
                val = str(cell.value or "")
                if val.startswith("="):
                    # 수식인 경우 대략적인 기본값 지정
                    length = 12
                else:
                    # 한글 등 유니코드 고려 대략적인 길이
                    length = sum(2 if ord(char) > 128 else 1 for char in val)
                if length > max_len:
                    max_len = length
            # 최소 너비 10, 최대 45로 고정
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)
            
    # 대시보드 시트 특정 열 너비 추가 보정
    ws_dash.column_dimensions['A'].width = 20
    ws_dash.column_dimensions['F'].width = 16
    
    # 저장
    wb.save(excel_output_path)
    print(f"[성공] 엑셀 분석 대시보드가 {excel_output_path}에 생성되었습니다.")

if __name__ == "__main__":
    build_dashboard()
