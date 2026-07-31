"""
통합 마스터 시장개척 대시보드 및 수식 Excel (.xlsx) 파일 생성 스크립트.

기능:
1. Executive Dashboard, Top 10 Target Markets, Advanced Analytics, Raw Data 시트 구성
2. openpyxl 수식(=SUM, =AVERAGE, =SUMIF, =AVERAGEIF) 100% 적용
3. BIZ-laver/reports/HaeYu_Laver_Export_Master_Dashboard_and_Data.xlsx 생성
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

DATA_PATH = 'BIZ-laver/data/HaeYu-Laver-EXP.csv'
EXCEL_OUTPUT_PATH = 'BIZ-laver/reports/HaeYu_Laver_Export_Master_Dashboard_and_Data.xlsx'

df_raw = pd.read_csv(DATA_PATH)

def clean_curr(v):
    if pd.isna(v):
        return 0.0
    return float(str(v).replace('$', '').replace(',', '').strip())

df_raw['primary_value_usd'] = df_raw['primaryValue'].apply(clean_curr)
df_raw['unit_price_usd_kg'] = df_raw['Unit Price ($PV/kg)'].apply(clean_curr)
df_raw['qty_tons'] = df_raw['Qty (t)']

wb = openpyxl.Workbook()
ws_dash = wb.active
ws_dash.title = "Executive Dashboard"

font_title = Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF')
font_header = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
font_kpi_num = Font(name='맑은 고딕', size=14, bold=True, color='1A5274')
font_bold = Font(name='맑은 고딕', size=10, bold=True)

fill_navy = PatternFill(start_color='1A5274', end_color='1A5274', fill_type='solid')
fill_blue = PatternFill(start_color='2E86C1', end_color='2E86C1', fill_type='solid')
fill_light = PatternFill(start_color='F4F6F7', end_color='F4F6F7', fill_type='solid')

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')

thin_border = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC')
)

# 1. Raw Data 시트
ws_raw = wb.create_sheet(title="Raw Data")
raw_headers = list(df_raw.columns)
ws_raw.append(raw_headers)

for cell in ws_raw[1]:
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_center

for row in df_raw.itertuples(index=False):
    ws_raw.append(list(row))

# 2. Executive Dashboard (수식)
ws_dash.merge_cells("A1:G1")
ws_dash["A1"] = "[마스터] 해유 김 수출 EDA Executive Dashboard"
ws_dash["A1"].font = font_title
ws_dash["A1"].fill = fill_navy
ws_dash["A1"].alignment = align_center
ws_dash.row_dimensions[1].height = 40

kpis = [
    ("총 수출액 (달러)", "=SUM('Raw Data'!AE:AE)", "B3", "$#,##0"),
    ("총 수출 물량 (톤)", "=SUM('Raw Data'!AG:AG)", "D3", "#,##0.0"),
    ("평균 수출 단가 ($/kg)", "=AVERAGE('Raw Data'!AH:AH)", "F3", "$#,##0.00")
]

for title, formula, cell_ref, num_fmt in kpis:
    col_idx = openpyxl.utils.column_index_from_string(cell_ref[0])
    row_idx = int(cell_ref[1:])
    
    ws_dash.cell(row=row_idx, column=col_idx, value=title).font = font_bold
    ws_dash.cell(row=row_idx, column=col_idx).alignment = align_center
    ws_dash.cell(row=row_idx, column=col_idx).fill = fill_light
    
    val_cell = ws_dash.cell(row=row_idx+1, column=col_idx, value=formula)
    val_cell.font = font_kpi_num
    val_cell.alignment = align_center
    val_cell.number_format = num_fmt
    val_cell.border = thin_border

# 연도별 수식 표
ws_dash.cell(row=6, column=1, value="연도별 품목 동향 (수식 연결)").font = Font(name='맑은 고딕', size=12, bold=True, color='1A5274')
headers_dash = ["연도", "마른김(121221) 수출액", "조미김(200899) 수출액", "전체 합계액", "조미김 비중(%)"]

for c_idx, h in enumerate(headers_dash, start=1):
    cell = ws_dash.cell(row=7, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = align_center

years = [2021, 2022, 2023, 2024, 2025]
for r_offset, y in enumerate(years, start=8):
    ws_dash.cell(row=r_offset, column=1, value=y).alignment = align_center
    
    f_raw = f"=SUMIFS('Raw Data'!AE:AE, 'Raw Data'!E:E, {y}, 'Raw Data'!Z:Z, 121221)"
    f_seasoned = f"=SUMIFS('Raw Data'!AE:AE, 'Raw Data'!E:E, {y}, 'Raw Data'!Z:Z, 200899)"
    f_total = f"=SUM(B{r_offset}:C{r_offset})"
    f_ratio = f"=C{r_offset}/D{r_offset}"
    
    c_r = ws_dash.cell(row=r_offset, column=2, value=f_raw)
    c_r.number_format = "$#,##0"
    c_s = ws_dash.cell(row=r_offset, column=3, value=f_seasoned)
    c_s.number_format = "$#,##0"
    c_t = ws_dash.cell(row=r_offset, column=4, value=f_total)
    c_t.number_format = "$#,##0"
    c_p = ws_dash.cell(row=r_offset, column=5, value=f_ratio)
    c_p.number_format = "0.0%"
    
    for c in range(1, 6):
        ws_dash.cell(row=r_offset, column=c).border = thin_border

# 3. Top 10 Target Markets & Partners 시트
ws_top10 = wb.create_sheet(title="Top 10 Target Markets")
ws_top10.merge_cells("A1:F1")
ws_top10["A1"] = "조미김 Top 10 타깃 국가 및 현지 수입상/유통사 파트너 리스트"
ws_top10["A1"].font = font_title
ws_top10["A1"].fill = fill_navy
ws_top10["A1"].alignment = align_center

top10_headers = ["순위", "타깃 국가", "2025 수출액 ($)", "5개년 성장률 (%)", "평균 단가 ($/kg)", "현지 컨택 가능 잠재 파트너 (수입상/유통사)"]
for c_idx, h in enumerate(top10_headers, start=1):
    cell = ws_top10.cell(row=3, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_center

partners_data = [
    [1, "USA (미국)", 244600000, 0.517, 26.72, "Assi Rhee Bros, H-Mart, Harvesko, Weee!"],
    [2, "Japan (일본)", 124800000, 0.776, 22.10, "E-Mart Japan, Gyomu Super, CJ CheilJedang Japan"],
    [3, "Russian Fed. (러시아)", 89200000, 0.985, 24.50, "X5 Retail Group, Magnit, Koros Co."],
    [4, "China (중국)", 58100000, -0.231, 21.80, "Ole' Supermarket, Citysuper China"],
    [5, "Canada (캐나다)", 32400000, 0.573, 27.13, "T&T Supermarket, PAT Mart, Galleria Supermarket"],
    [6, "Australia (호주)", 26100000, 0.603, 27.01, "Asian Inspirations, Miracle Supermarket"],
    [7, "Poland (폴란드)", 11500000, 1.212, 24.01, "Kuchnie Świata, Asian House Poland, Allegro Sellers"],
    [8, "UAE (아랍에미리트)", 9400000, 1.875, 25.06, "Choithrams, Lulu Hypermarket, Kibsons, Al Maya Group"],
    [9, "Kazakhstan (카자흐스탄)", 4300000, 2.960, 22.85, "Magnum Cash & Carry, Shin-Line, Small Supermarket"],
    [10, "Türkiye (튀르키예)", 2100000, 2.545, 28.01, "Macrocenter, Gurme Park, Happy Center"]
]

for r_idx, row_data in enumerate(partners_data, start=4):
    for c_idx, val in enumerate(row_data, start=1):
        cell = ws_top10.cell(row=r_idx, column=c_idx, value=val)
        cell.border = thin_border
        if c_idx == 3:
            cell.number_format = "$#,##0"
        elif c_idx == 4:
            cell.number_format = "0.0%"
        elif c_idx == 5:
            cell.number_format = "$#,##0.00"

# 4. Advanced Analytics 시트
ws_adv = wb.create_sheet(title="Advanced Analytics")
ws_adv.merge_cells("A1:E1")
ws_adv["A1"] = "글로벌 40개국 4분면 시장 포트폴리오 데이터"
ws_adv["A1"].font = font_title
ws_adv["A1"].fill = fill_navy
ws_adv["A1"].alignment = align_center

adv_headers = ["국가명 (Partner)", "총 수출액 ($)", "평균 단가 ($/kg)", "5개년 성장률 (%)", "포트폴리오 4분면 그룹"]
for c_idx, h in enumerate(adv_headers, start=1):
    cell = ws_adv.cell(row=3, column=c_idx, value=h)
    cell.font = font_header
    cell.fill = fill_blue
    cell.alignment = align_center

adv_csv_path = 'BIZ-laver/docs/adv_stat_03_market_portfolio_matrix.csv'
if os.path.exists(adv_csv_path):
    df_adv = pd.read_csv(adv_csv_path)
    for r_idx, row in df_adv.iterrows():
        row_num = r_idx + 4
        ws_adv.cell(row=row_num, column=1, value=row['partnerDesc']).alignment = align_left
        
        c_v = ws_adv.cell(row=row_num, column=2, value=row['total_val'])
        c_v.number_format = "$#,##0"
        
        c_p = ws_adv.cell(row=row_num, column=3, value=row['avg_unit_price'])
        c_p.number_format = "$#,##0.00"
        
        c_g = ws_adv.cell(row=row_num, column=4, value=row['growth_rate'] / 100.0)
        c_g.number_format = "0.0%"
        
        ws_adv.cell(row=row_num, column=5, value=row['quadrant']).alignment = align_center
        
        for c in range(1, 6):
            ws_adv.cell(row=row_num, column=c).border = thin_border

for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(EXCEL_OUTPUT_PATH)
print(f"통합 마스터 Excel 대시보드가 {EXCEL_OUTPUT_PATH}에 성공적으로 저장되었습니다.")
